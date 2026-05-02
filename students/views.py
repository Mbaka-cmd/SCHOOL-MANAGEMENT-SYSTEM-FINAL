from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Student
from academics.models import Stream, ClassLevel
from schools.models import AcademicYear
from schools.views import admin_required
from exams.models import ReportCard, Exam


@admin_required
def students_by_stream(request):
    """View students organized by streams with option to generate report cards"""
    school = request.user.school
    streams = Stream.objects.filter(school=school).order_by("class_level__level_order", "name")
    
    # Get selected stream from query parameters
    stream_id = request.GET.get("stream")
    selected_stream = None
    students_in_stream = []
    exams = []
    
    if stream_id:
        selected_stream = get_object_or_404(Stream, id=stream_id, school=school)
        students_in_stream = Student.objects.filter(
            school=school, 
            current_stream=selected_stream, 
            is_active=True
        ).order_by("last_name", "first_name").prefetch_related('parents')
        # Get available exams for report card generation
        exams = Exam.objects.filter(school=school, is_published=True).order_by("-created_at")
    
    context = {
        "streams": streams,
        "selected_stream": selected_stream,
        "students_in_stream": students_in_stream,
        "exams": exams,
    }
    return render(request, "students/students_by_stream.html", context)


@admin_required
def student_list(request):
    school = request.user.school
    students = Student.objects.filter(school=school, is_active=True)

    stream_id = request.GET.get("stream")
    class_level_id = request.GET.get("class_level")
    search = request.GET.get("search", "")

    if stream_id:
        students = students.filter(current_stream_id=stream_id)
    if class_level_id:
        students = students.filter(current_stream__class_level_id=class_level_id)
    if search:
        students = students.filter(
            first_name__icontains=search
        ) | students.filter(
            last_name__icontains=search
        ) | students.filter(
            admission_number__icontains=search
        )

    class_levels = ClassLevel.objects.filter(school=school)
    streams = Stream.objects.filter(school=school)

    context = {
        "students": students.order_by("current_stream__class_level__level_order", "last_name"),
        "class_levels": class_levels,
        "streams": streams,
        "total": students.count(),
        "search": search,
    }
    return render(request, "students/student_list.html", context)


@admin_required
def student_add(request):
    school = request.user.school
    if request.method == "POST":
        try:
            student = Student(
                school=school,
                admission_number=request.POST.get("admission_number"),
                first_name=request.POST.get("first_name"),
                last_name=request.POST.get("last_name"),
                middle_name=request.POST.get("middle_name", ""),
                date_of_birth=request.POST.get("date_of_birth"),
                gender=request.POST.get("gender", "female"),
                admission_date=request.POST.get("admission_date"),
                admission_type=request.POST.get("admission_type", "form1"),
                is_boarder=request.POST.get("is_boarder") == "on",
                dormitory=request.POST.get("dormitory", ""),
                kcpe_marks=request.POST.get("kcpe_marks") or None,
                nemis_number=request.POST.get("nemis_number", ""),
            )
            stream_id = request.POST.get("current_stream")
            if stream_id:
                student.current_stream = Stream.objects.get(id=stream_id, school=school)
            student.save()
            messages.success(request, f"Student {student.get_full_name()} added successfully!")
            return redirect("student_list")
        except Exception as e:
            messages.error(request, f"Error adding student: {e}")

    streams = Stream.objects.filter(school=school).order_by("class_level__level_order", "name")
    context = {"streams": streams}
    return render(request, "students/student_add.html", context)


@admin_required
def student_detail(request, pk):
    school = request.user.school
    student = get_object_or_404(Student, id=pk, school=school)
    context = {"student": student}
    return render(request, "students/student_detail.html", context)


@admin_required
def student_edit(request, pk):
    school = request.user.school
    student = get_object_or_404(Student, id=pk, school=school)
    if request.method == "POST":
        student.first_name = request.POST.get("first_name")
        student.last_name = request.POST.get("last_name")
        student.middle_name = request.POST.get("middle_name", "")
        student.date_of_birth = request.POST.get("date_of_birth")
        student.is_boarder = request.POST.get("is_boarder") == "on"
        student.dormitory = request.POST.get("dormitory", "")
        student.nemis_number = request.POST.get("nemis_number", "")
        student.kcpe_marks = request.POST.get("kcpe_marks") or None
        stream_id = request.POST.get("current_stream")
        if stream_id:
            student.current_stream = Stream.objects.get(id=stream_id, school=school)
        student.save()
        messages.success(request, "Student updated successfully!")
        return redirect("student_detail", pk=student.id)

    streams = Stream.objects.filter(school=school).order_by("class_level__level_order", "name")
    context = {"student": student, "streams": streams}
    return render(request, "students/student_edit.html", context)


def resolve_stream(school, stream_name):
    """Resolve stream name like 'Form 1 East' to Stream object"""
    if not stream_name:
        return None
    parts = stream_name.strip().split()
    # Handle "Form 1 East" format (3+ parts)
    if len(parts) >= 3:
        level_name = ' '.join(parts[:2])  # "Form 1"
        name = parts[2]                    # "East"
        stream = Stream.objects.filter(
            school=school,
            name__iexact=name,
            class_level__name__iexact=level_name
        ).first()
        if stream:
            return stream
    # Fallback: search by stream name only
    stream = Stream.objects.filter(
        school=school,
        name__icontains=stream_name
    ).first()
    if stream:
        return stream
    # Fallback: search by class level name
    stream = Stream.objects.filter(
        school=school,
        class_level__name__icontains=stream_name
    ).first()
    return stream


@admin_required
def student_bulk_import(request):
    """Bulk import students from Excel file"""
    school = request.user.school
    streams = Stream.objects.filter(school=school).order_by('class_level__level_order', 'name')

    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "Please upload an Excel file.")
            return render(request, 'students/bulk_import.html', {'streams': streams})

        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, "Please upload a valid Excel file (.xlsx or .xls).")
            return render(request, 'students/bulk_import.html', {'streams': streams})

        try:
            import openpyxl
            from datetime import datetime

            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            created = 0
            skipped = 0
            errors = []

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue

                try:
                    # --- SAFE COLUMN ACCESS (handles missing columns) ---
                    def get_val(index, default=None):
                        """Safely get value by index, return default if out of range or empty"""
                        if index >= len(row):
                            return default
                        val = row[index]
                        if val is None or val == '':
                            return default
                        return val

                    admission_number = str(get_val(0)).strip() if get_val(0) else None
                    first_name = str(get_val(1)).strip() if get_val(1) else None
                    last_name = str(get_val(2)).strip() if get_val(2) else None
                    middle_name = str(get_val(3, '')).strip()
                    gender = str(get_val(4, 'female')).strip().lower()
                    dob = get_val(5)
                    stream_name = str(get_val(6)).strip() if get_val(6) else None
                    kcpe_marks = get_val(7)
                    admission_date = get_val(8)
                    is_boarder_val = get_val(9)
                    is_boarder = str(is_boarder_val).strip().lower() in ['yes', 'true', '1'] if is_boarder_val is not None else False

                    if not admission_number or not first_name or not last_name:
                        errors.append(f"Row {row_num}: Missing required fields")
                        skipped += 1
                        continue

                    if Student.objects.filter(admission_number=admission_number, school=school).exists():
                        errors.append(f"Row {row_num}: Admission number {admission_number} already exists")
                        skipped += 1
                        continue

                    stream = resolve_stream(school, stream_name)

                    if isinstance(dob, str):
                        try:
                            dob = datetime.strptime(dob, '%Y-%m-%d').date()
                        except:
                            dob = None
                    elif hasattr(dob, 'date'):
                        dob = dob.date()

                    if isinstance(admission_date, str):
                        try:
                            admission_date = datetime.strptime(admission_date, '%Y-%m-%d').date()
                        except:
                            admission_date = None
                    elif hasattr(admission_date, 'date'):
                        admission_date = admission_date.date()

                    if gender not in ['male', 'female']:
                        gender = 'female'

                    Student.objects.create(
                        school=school,
                        admission_number=admission_number,
                        first_name=first_name,
                        last_name=last_name,
                        middle_name=middle_name,
                        gender=gender,
                        date_of_birth=dob,
                        current_stream=stream,
                        kcpe_marks=kcpe_marks,
                        admission_date=admission_date,
                        is_boarder=is_boarder,
                        is_active=True,
                    )
                    created += 1

                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    skipped += 1

            if created > 0:
                messages.success(request, f"Successfully imported {created} students!")
            if skipped > 0:
                messages.warning(request, f"Skipped {skipped} rows. Check errors below.")

            context = {
                'streams': streams,
                'created': created,
                'skipped': skipped,
                'errors': errors,
                'done': True,
            }
            return render(request, 'students/bulk_import.html', context)

        except Exception as e:
            messages.error(request, f"Error reading file: {str(e)}")

    return render(request, 'students/bulk_import.html', {'streams': streams})


@admin_required
def download_import_template(request):
    """Download Excel template for bulk import"""
    import openpyxl
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students Import"

    headers = [
        'admission_number', 'first_name', 'last_name', 'middle_name',
        'gender', 'date_of_birth', 'stream', 'kcpe_marks',
        'admission_date', 'is_boarder'
    ]
    ws.append(headers)

    # Sample row
    ws.append([
        'CGS001/2026', 'Jane', 'Doe', 'Mary',
        'female', '2010-01-15', 'Form 1 East', 380,
        '2026-01-10', 'yes'
    ])

    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="students_import_template.xlsx"'
    wb.save(response)
    return response